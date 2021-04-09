# импорт модулей openpyxl и tkinter

from openpyxl import *

from tkinter import *

  
# глобально объявляем переменную wb и sheet

  
# открытие существующего файла excel

wb = load_workbook('C:\\Users\\user\\Desktop\\excel.xlsx')

  
# создать объект листа

sheet = wb.active

  

  

def excel():

      

    # изменить ширину столбцов в

    # таблица Excel

    sheet.column_dimensions['A'].width = 30

    sheet.column_dimensions['B'].width = 10

    sheet.column_dimensions['C'].width = 10

    sheet.column_dimensions['D'].width = 20

    sheet.column_dimensions['E'].width = 20

    sheet.column_dimensions['F'].width = 40

    sheet.column_dimensions['G'].width = 50

  

    # записать данные в таблицу Excel

    # в определенном месте

    sheet.cell(row=1, column=1).value = "Name"

    sheet.cell(row=1, column=2).value = "Course"

    sheet.cell(row=1, column=3).value = "Semester"

    sheet.cell(row=1, column=4).value = "Form Number"

    sheet.cell(row=1, column=5).value = "Contact Number"

    sheet.cell(row=1, column=6).value = "Email id"

    sheet.cell(row=1, column=7).value = "Address"

  

  
# Функция для установки фокуса (курсор)

def focus1(event):

    # установить фокус на поле course_field

    course_field.focus_set()

  

  
# Функция для установки фокуса

def focus2(event):

    # установить фокус на поле sem_field

    sem_field.focus_set()

  

  
# Функция для установки фокуса

def focus3(event):

    # установить фокус на поле form_no_field

    form_no_field.focus_set()

  

  
# Функция для установки фокуса

def focus4(event):

    # установить фокус на поле contact_no_field

    contact_no_field.focus_set()

  

  
# Функция для установки фокуса

def focus5(event):

    # установить фокус на поле email_id_field

    email_id_field.focus_set()

  

  
# Функция для установки фокуса

def focus6(event):

    # установить фокус на поле address_field

    address_field.focus_set()

  

  
# Функция для очистки
# содержимое полей ввода текста

def clear():

      

    # очистить содержимое поля ввода текста

    name_field.delete(0, END)

    course_field.delete(0, END)

    sem_field.delete(0, END)

    form_no_field.delete(0, END)

    contact_no_field.delete(0, END)

    email_id_field.delete(0, END)

    address_field.delete(0, END)

  

  
# Функция для получения данных из GUI
# Окно и запись в файл Excel

def insert():

      

    # если пользователь не заполняет какую-либо запись

    # затем выведите «пустой ввод»

    if (name_field.get() == "" and

        course_field.get() == "" and

        sem_field.get() == "" and

        form_no_field.get() == "" and

        contact_no_field.get() == "" and

        email_id_field.get() == "" and

        address_field.get() == ""):

              

        print("empty input")

  

    else:

  

        # назначение строки макс и столбца макс

        # значение, до которого записываются данные

        # в листе Excel к переменной

        current_row = sheet.max_row

        current_column = sheet.max_column

  

        # метод get возвращает текущий текст

        # как строка, в которую мы записываем

        # таблица Excel в определенном месте

        sheet.cell(row=current_row + 1, column=1).value = name_field.get()

        sheet.cell(row=current_row + 1, column=2).value = course_field.get()

        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()

        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()

        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()

        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()

        sheet.cell(row=current_row + 1, column=7).value = address_field.get()

  

        # сохранить файл

        wb.save('C:\\Users\\user\\Desktop\\excel.xlsx')

  

        # установить фокус на поле name_field

        name_field.focus_set()

  

        # вызвать функцию clear ()

        clear()

  

  
# Код драйвера

if __name__ == "__main__":

      

    # создать окно с графическим интерфейсом

    root = Tk()

  

    # установить цвет фона окна GUI

    root.configure(background='light green')

  

    # установить заголовок окна GUI

    root.title("registration form")

  

    # установить конфигурацию окна графического интерфейса

    root.geometry("500x300")

  

    excel()

  

    # создать метку формы

    heading = Label(root, text="Form", bg="light green")

  

    # создать метку имени

    name = Label(root, text="Name", bg="light green")

  

    # создать ярлык курса

    course = Label(root, text="Course", bg="light green")

  

    # создать семестр

    sem = Label(root, text="Semester", bg="light green")

  

    # создать форму № метки

    form_no = Label(root, text="Form No.", bg="light green")

  

    # создать метку № контакта

    contact_no = Label(root, text="Contact No.", bg="light green")

  

    # создать ярлык с идентификатором

    email_id = Label(root, text="Email id", bg="light green")

  

    # создать адресную метку

    address = Label(root, text="Address", bg="light green")

  

    # метод сетки используется для размещения

    # виджеты на соответствующих позициях

    # в таблице как структура.

    heading.grid(row=0, column=1)

    name.grid(row=1, column=0)

    course.grid(row=2, column=0)

    sem.grid(row=3, column=0)

    form_no.grid(row=4, column=0)

    contact_no.grid(row=5, column=0)

    email_id.grid(row=6, column=0)

    address.grid(row=7, column=0)

  

    # создать текстовое поле ввода

    # для ввода информации

    name_field = Entry(root)

    course_field = Entry(root)

    sem_field = Entry(root)

    form_no_field = Entry(root)

    contact_no_field = Entry(root)

    email_id_field = Entry(root)

    address_field = Entry(root)

  

    # метод привязки виджета используется для

    # связывание функции с событиями

  

    # при каждом нажатии клавиши ввода

    # затем вызовите функцию focus1

    name_field.bind("<Return>", focus1)

  

    # при каждом нажатии клавиши ввода

    # затем вызовите функцию focus2

    course_field.bind("<Return>", focus2)

  

    # при каждом нажатии клавиши ввода

    # затем вызовите функцию focus3

    sem_field.bind("<Return>", focus3)

  

    # при каждом нажатии клавиши ввода

    # затем вызовите функцию focus4

    form_no_field.bind("<Return>", focus4)

  

    # при каждом нажатии клавиши ввода

    # затем вызовите функцию focus5

    contact_no_field.bind("<Return>", focus5)

  

    # при каждом нажатии клавиши ввода

    # затем вызовите функцию focus6

    email_id_field.bind("<Return>", focus6)

  

    # метод сетки используется для размещения

    # виджеты на соответствующих позициях

    # в таблице как структура.

    name_field.grid(row=1, column=1, ipadx="100")

    course_field.grid(row=2, column=1, ipadx="100")

    sem_field.grid(row=3, column=1, ipadx="100")

    form_no_field.grid(row=4, column=1, ipadx="100")

    contact_no_field.grid(row=5, column=1, ipadx="100")

    email_id_field.grid(row=6, column=1, ipadx="100")

    address_field.grid(row=7, column=1, ipadx="100")

  

    # вызов функции Excel

    excel()

  

    # создать кнопку отправки и поместить в корневое окно

    submit = Button(root, text="Submit", fg="Black",

                            bg="Red", command=insert)

    submit.grid(row=8, column=1)

  

    # запустить графический интерфейс

    root.mainloop() 